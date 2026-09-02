#!/usr/bin/env python3
"""Build wine_auction_prices.sqlite from scratch (repeatable, deterministic).

Reads all structured auction data (14 JSON houses in the Google Drive archive
plus the Wermuth and Steinfels Excel/CSV in the winefeed repo) and writes a
single normalized SQLite database.

Every run drops and recreates the schema, so the result is fully reproducible:
    python3 build_db.py            # full build
    python3 build_db.py --limit 20 # dry run on a few files per house
    python3 build_db.py --dry-run  # no write
"""
import argparse
import glob
import json
import os
import re
import sqlite3
import sys

import openpyxl

# Source of raw auction data (Google Drive archive). Override via env var:
#   ARCHIVE_PATH=/path/to/archive python3 build_db.py
ARCHIVE = os.environ.get("ARCHIVE_PATH", os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-persi.marco@gmail.com/"
    "Meine Ablage/Wein/WeinAuktionspreise"))

# Repo root (contains the Wermuth/Steinfels Excel files) and DB output location.
# Both are derived from this script's location (data/ inside the repo).
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(SCRIPT_DIR)
DB_PATH = os.path.join(SCRIPT_DIR, "wine_auction_prices.sqlite")

# --------------------------------------------------------------------------- #
# Schema
# --------------------------------------------------------------------------- #

SCHEMA = """
DROP TABLE IF EXISTS wine_alias_resolutions;
DROP TABLE IF EXISTS wine_alias_observations;
DROP TABLE IF EXISTS build_metrics;
DROP TABLE IF EXISTS build_inputs;
DROP TABLE IF EXISTS builds;
DROP TABLE IF EXISTS lots;
DROP TABLE IF EXISTS wines;
DROP TABLE IF EXISTS auctions;
DROP TABLE IF EXISTS providers;

CREATE TABLE providers (
  id   INTEGER PRIMARY KEY,
  name TEXT NOT NULL UNIQUE
);

CREATE TABLE auctions (
  id          INTEGER PRIMARY KEY,
  provider_id INTEGER NOT NULL REFERENCES providers(id),
  auction_id  TEXT,
  title       TEXT,
  auction_date TEXT,
  currency    TEXT,
  price_basis TEXT,
  source_url  TEXT,
  scraped_at  TEXT,
  UNIQUE(provider_id, auction_id)
);

CREATE TABLE wines (
  id               INTEGER PRIMARY KEY,
  stable_key       TEXT NOT NULL UNIQUE,
  canonical_name   TEXT NOT NULL,
  canonical_key    TEXT NOT NULL,
  producer         TEXT,
  producer_key     TEXT,
  region           TEXT,
  appellation      TEXT,
  classification   TEXT,
  canonical_source TEXT NOT NULL,
  resolver_version TEXT NOT NULL,
  review_status    TEXT NOT NULL,
  CHECK (trim(canonical_name) <> ''),
  CHECK (review_status IN (
    'AUTO_RESOLVED', 'CURATED', 'AMBIGUOUS', 'REVIEW_REQUIRED'
  ))
);

CREATE TABLE lots (
  id                  INTEGER PRIMARY KEY,
  auction_id          INTEGER NOT NULL REFERENCES auctions(id),
  lot_no              TEXT,
  lot_date            TEXT,
  source_file         TEXT,
  source_file_sha256  TEXT,
  source_record_locator TEXT,
  source_record_hash  TEXT,
  raw_wine            TEXT,
  raw_producer        TEXT,
  raw_vintage         TEXT,
  source_lot_key      TEXT,
  wine                TEXT,
  wine_ref_id         INTEGER REFERENCES wines(id),
  producer            TEXT,
  vintage             INTEGER,
  region              TEXT,
  classification      TEXT,
  vintage_raw         TEXT,
  vintage_extracted   INTEGER,
  vintage_final       INTEGER,
  vintage_status      TEXT,
  vintage_rule_id     TEXT,
  vintage_evidence_text TEXT,
  lot_kind            TEXT,
  mixed_reason        TEXT,
  mixed_rule_id       TEXT,
  mixed_lot           INTEGER,
  quantity            INTEGER,
  bottle_size_dl      REAL,
  estimate_low        REAL,
  estimate_high       REAL,
  hammer_price        REAL,
  buyer_premium       REAL,
  realised_price      REAL,
  price_basis         TEXT,
  currency            TEXT,
  sold                INTEGER,
  condition_text      TEXT,
  description         TEXT,
  source_url          TEXT,
  scraped_at          TEXT,
  CHECK (vintage_final IS NULL OR vintage_final BETWEEN 1700 AND 2100),
  CHECK (lot_kind IS NULL OR lot_kind IN ('SINGLE', 'MIXED', 'UNKNOWN')),
  CHECK (lot_kind <> 'MIXED' OR wine_ref_id IS NULL),
  CHECK (vintage_status IS NULL OR vintage_status IN (
    'STRUCTURED', 'EXTRACTED', 'NV', 'MV', 'MISSING',
    'AMBIGUOUS', 'CONFLICT', 'INVALID'
  ))
);

CREATE TABLE wine_alias_observations (
  id                 INTEGER PRIMARY KEY,
  provider_id        INTEGER NOT NULL REFERENCES providers(id),
  alias_raw          TEXT NOT NULL,
  alias_key          TEXT NOT NULL,
  producer_raw       TEXT,
  producer_key       TEXT,
  region_raw         TEXT,
  region_key         TEXT,
  source_lot_id      INTEGER REFERENCES lots(id),
  evidence_count     INTEGER NOT NULL,
  normalizer_version TEXT NOT NULL
);

CREATE TABLE wine_alias_resolutions (
  id                INTEGER PRIMARY KEY,
  scope_key         TEXT NOT NULL,
  alias_key         TEXT NOT NULL,
  discriminator_key TEXT NOT NULL,
  wine_id           INTEGER REFERENCES wines(id),
  resolution_status TEXT NOT NULL,
  resolution_method TEXT NOT NULL,
  confidence        REAL,
  rule_version      TEXT NOT NULL,
  reason            TEXT,
  UNIQUE(scope_key, alias_key, discriminator_key),
  CHECK (resolution_status IN (
    'RESOLVED', 'AMBIGUOUS', 'UNRESOLVED', 'IGNORED'
  ))
);

CREATE TABLE builds (
  id                    TEXT PRIMARY KEY,
  started_at            TEXT NOT NULL,
  completed_at          TEXT,
  input_manifest_sha256 TEXT NOT NULL,
  loader_version        TEXT NOT NULL,
  normalizer_version    TEXT NOT NULL,
  resolver_version      TEXT NOT NULL,
  python_version        TEXT NOT NULL,
  sqlite_version        TEXT NOT NULL,
  status                TEXT NOT NULL
);

CREATE TABLE build_inputs (
  build_id         TEXT NOT NULL REFERENCES builds(id),
  source_file      TEXT NOT NULL,
  source_file_sha256 TEXT NOT NULL,
  file_size        INTEGER NOT NULL,
  PRIMARY KEY (build_id, source_file)
);

CREATE TABLE build_metrics (
  build_id TEXT NOT NULL REFERENCES builds(id),
  provider TEXT NOT NULL,
  metric   TEXT NOT NULL,
  value    INTEGER NOT NULL,
  PRIMARY KEY (build_id, provider, metric)
);

CREATE INDEX idx_lots_auction ON lots(auction_id);
CREATE INDEX idx_lots_wine ON lots(wine);
CREATE INDEX idx_lots_hammer ON lots(hammer_price);
CREATE INDEX idx_lots_wine_ref ON lots(wine_ref_id);
CREATE INDEX idx_lots_lot_kind ON lots(lot_kind);

CREATE VIEW single_wine_prices AS
SELECT *
FROM lots
WHERE lot_kind = 'SINGLE'
  AND wine_ref_id IS NOT NULL
  AND vintage_status IN ('STRUCTURED', 'EXTRACTED');
"""

# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"])}


def load_json(path):
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def to_float(v):
    """Best-effort numeric conversion (US/UK thousands separator)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dict):
        return None
    s = str(v).strip()
    s = s.replace("$", "").replace("€", "").replace("£", "")
    s = s.replace("HK$", "").replace("CHF", "").replace("\u00a0", "")
    s = s.replace(" ", "")
    if s in ("", "-", "null", "None", "—"):
        return None
    if "," in s and "." in s:
        s = s.replace(",", "")
    elif "," in s:
        # "1,434" -> thousands ; "1,5" -> decimal
        if re.fullmatch(r"\d{1,3},\d{3}(,\d{3})*", s):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def euro_int(v):
    """European thousands format: '8.060' -> 8060, '10.000' -> 10000."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, dict):
        return to_float(v.get("value"))
    s = str(v).strip().replace("€", "").replace(" ", "").replace("\u00a0", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def to_int(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, bool):
        return None
    s = str(v).strip()
    if s in ("", "null", "None"):
        return None
    m = re.search(r"-?\d+", s.replace(",", ""))
    return int(m.group()) if m else None


def parse_date(v):
    if not v:
        return None
    if isinstance(v, dict):
        for k in ("closed", "acceptsBids", "startsToClose", "published"):
            if v.get(k):
                v = v[k]
                break
        else:
            return None
    s = str(v).strip()
    if s.startswith("20") and len(s) >= 10 and s[4] == "-":
        return s[:10]
    m = re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{1,2})\s+(\w{3,9})\.?\s+(\d{4})", s)
    if m:
        mon = MONTHS.get(m.group(2).lower()[:3])
        if mon:
            return f"{m.group(3)}-{mon:02d}-{int(m.group(1)):02d}"
    return None


def ym_date(year, month):
    """Year+month -> ISO-ish date (first of month)."""
    if year is None:
        return None
    if month is None:
        return f"{int(year)}-01-01"
    return f"{int(year)}-{int(month):02d}-01"


def parse_dl(v):
    """Bottle size hints -> deciliters (7.5 = 750 ml)."""
    if v is None or v == "":
        return None
    s = str(v).lower()
    m = re.search(r"(\d+(?:\.\d+)?)\s*(ml|l)", s)
    if m:
        n = float(m.group(1))
        return n / 100 if m.group(2) == "ml" else n * 10
    if "magnum" in s or "1.5" in s or "1500" in s:
        return 15.0
    if "jeroboam" in s or "3l" in s or "3000ml" in s:
        return 30.0
    if "half" in s or "375" in s or "3.75" in s:
        return 3.75
    f = to_float(v)
    if f is not None and 1 <= f <= 60:
        return f
    return None


def parse_qty_text(v):
    """'2 Bottles 750ml' -> 2."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"(\d+)", str(v))
    return int(m.group(1)) if m else None


# --------------------------------------------------------------------------- #
# Insert helpers
# --------------------------------------------------------------------------- #

class Builder:
    def __init__(self, conn):
        self.conn = conn
        self.provider_ids = {}
        self.source_file = None
        self.source_idx = 0

    def provider(self, name):
        if name not in self.provider_ids:
            cur = self.conn.execute(
                "INSERT INTO providers(name) VALUES (?)", (name,))
            self.provider_ids[name] = cur.lastrowid
        return self.provider_ids[name]

    def set_source(self, source_file):
        """Set provenance context for the following add_lot calls."""
        self.source_file = os.path.relpath(source_file, ARCHIVE)
        self.source_idx = 0

    def add_auction(self, provider, auction_id, title=None, date=None,
                    currency=None, price_basis=None, source_url=None,
                    scraped_at=None):
        pid = self.provider(provider)
        cur = self.conn.execute(
            "INSERT OR IGNORE INTO auctions(provider_id, auction_id, title,"
            " auction_date, currency, price_basis, source_url, scraped_at)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (pid, auction_id, title, date, currency, price_basis,
             source_url, scraped_at))
        if cur.rowcount == 1:
            return cur.lastrowid
        row = self.conn.execute(
            "SELECT id FROM auctions WHERE provider_id=? AND auction_id=?",
            (pid, auction_id)).fetchone()
        return row[0]

    def add_lot(self, auction_id, row):
        wine = row.get("wine")
        producer = row.get("producer")
        vintage = row.get("vintage")
        raw_wine = row.get("raw_wine") if "raw_wine" in row else wine
        raw_producer = row.get("raw_producer") if "raw_producer" in row else producer
        raw_vintage = row.get("raw_vintage")
        if raw_vintage is None and vintage is not None:
            raw_vintage = str(vintage)
        source_file = row.get("source_file", self.source_file)
        source_record_locator = row.get("source_record_locator")
        if source_record_locator is None:
            source_record_locator = f"lots[{self.source_idx}]"
            self.source_idx += 1
        cur = self.conn.execute(
            "INSERT INTO lots(auction_id, lot_no, lot_date, source_file,"
            " source_record_locator, raw_wine, raw_producer, raw_vintage,"
            " source_lot_key, wine, wine_ref_id, producer, vintage, region,"
            " classification, quantity, bottle_size_dl, estimate_low,"
            " estimate_high, hammer_price, buyer_premium, realised_price,"
            " price_basis, currency, sold, mixed_lot, condition_text,"
            " description, source_url, scraped_at)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (auction_id, row.get("lot_no"), row.get("lot_date"),
             source_file, source_record_locator,
             raw_wine, raw_producer, raw_vintage,
             row.get("source_lot_key"), wine, row.get("wine_ref_id"),
             producer, vintage, row.get("region"),
             row.get("classification"), row.get("quantity"),
             row.get("bottle_size_dl"), row.get("estimate_low"),
             row.get("estimate_high"), row.get("hammer_price"),
             row.get("buyer_premium"), row.get("realised_price"),
             row.get("price_basis"), row.get("currency"), row.get("sold"),
             row.get("mixed_lot"), row.get("condition_text"),
             row.get("description"), row.get("source_url"),
             row.get("scraped_at")))
        return cur.lastrowid


# --------------------------------------------------------------------------- #
# Per-house loaders
# --------------------------------------------------------------------------- #

def _json_files(house, pattern=None):
    base = os.path.join(ARCHIVE, house)
    pat = pattern or "**/*.json"
    files = sorted(glob.glob(os.path.join(base, pat), recursive=True))
    return [f for f in files if not f.endswith("auctions_directory.json")
            and not f.endswith("wines_directory.json")]


def load_baghera(b, limit):
    n = 0
    for f in _json_files("Baghera", "**/auction-*.json"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        cur = a.get("currency") or "CHF"
        a_id = b.add_auction(
            "baghera", aid, a.get("title"), ym_date(a.get("year"), a.get("month")),
            cur, "HAMMER", a.get("results_pdf"))
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": None, "description": lot.get("description"),
                "hammer_price": to_float(lot.get("hammer")),
                "realised_price": to_float(lot.get("total_price")),
                "buyer_premium": to_float(lot.get("buyer_premium")),
                "price_basis": "HAMMER", "currency": lot.get("currency") or cur,
            })
        n += 1
        if limit and n >= limit:
            break


def load_hdh(b, limit):
    n = 0
    for f in _json_files("HDH"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("sale_number"))
        a_id = b.add_auction(
            "hdh", aid, a.get("name"), ym_date(a.get("year"), a.get("month")),
            "USD", "HAMMER")
        for lot in d.get("lots", []):
            hammer = to_float(lot.get("hammer"))
            aggregate = to_float(lot.get("aggregate"))
            if hammer is not None:
                basis = "HAMMER"
            elif aggregate is not None:
                basis = "HAMMER_PLUS_BUYERS_PREMIUM"
            else:
                basis = "UNKNOWN"
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": lot.get("wine"), "producer": lot.get("grower"),
                "vintage": to_int(lot.get("vintage")),
                "quantity": to_int(lot.get("qty")),
                "bottle_size_dl": parse_dl(lot.get("size")),
                "estimate_low": to_float(lot.get("estimate_low")),
                "estimate_high": to_float(lot.get("estimate_high")),
                "hammer_price": hammer,
                "realised_price": aggregate,
                "price_basis": basis, "currency": "USD",
                "mixed_lot": 1 if lot.get("mixed_lot") else None,
                "condition_text": lot.get("condition"),
                "description": lot.get("description"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_winefields(b, limit):
    n = 0
    for f in _json_files("Winefields"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("auction_id"))
        a_id = b.add_auction(
            "winefields", aid, a.get("title"),
            parse_date(a.get("time_start")), None, "HAMMER")
        for lot in d.get("lots", []):
            cur = lot.get("currency")
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": lot.get("title"), "producer": lot.get("producer"),
                "quantity": to_int(lot.get("quantity")),
                "bottle_size_dl": parse_dl(lot.get("size")),
                "estimate_low": to_float(lot.get("estimate_low")),
                "estimate_high": to_float(lot.get("estimate_high")),
                "hammer_price": to_float(lot.get("hammer")),
                "price_basis": "HAMMER", "currency": cur,
                "sold": 1 if lot.get("status") == "sold" else 0,
                "mixed_lot": 1 if lot.get("is_mixed_lot") else 0,
                "condition_text": lot.get("condition"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_beschcannes(b, limit):
    n = 0
    for f in _json_files("BeschCannes", "**/auction-*.json"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        cur = a.get("currency") or "EUR"
        a_id = b.add_auction(
            "besch-cannes", aid, a.get("slug"), parse_date(a.get("date")),
            cur, a.get("price_basis") or "HAMMER")
        for lot in d.get("lots", []):
            sold = 1 if lot.get("adjuge") not in (None, "") else 0
            est = lot.get("estimate")
            est_low = est[0] if isinstance(est, list) and len(est) > 0 else None
            est_high = est[1] if isinstance(est, list) and len(est) > 1 else None
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": lot.get("appellation"),
                "vintage": to_int(lot.get("vintage")),
                "estimate_low": to_float(est_low),
                "estimate_high": to_float(est_high),
                "hammer_price": to_float(lot.get("hammer") or lot.get("adjuge")),
                "price_basis": "HAMMER", "currency": cur, "sold": sold,
                "description": lot.get("description"),
                "condition_text": lot.get("remarks"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_langtons(b, limit):
    n = 0
    for f in _json_files("Langtons"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        a_id = b.add_auction(
            "langtons", aid, a.get("name"), parse_date(a.get("date")),
            "AUD", "HAMMER")
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot")),
                "wine": lot.get("wine"), "region": lot.get("region"),
                "classification": lot.get("classification") or None,
                "vintage": to_int(lot.get("vintage")),
                "quantity": to_int(lot.get("quantity")),
                "hammer_price": to_float(lot.get("winning_bid")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency") or "AUD",
                "condition_text": lot.get("notes"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_sothebys(b, limit):
    n = 0
    for f in _json_files("Sothebys"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("auction_id"))
        cur = a.get("currency")
        a_id = b.add_auction(
            "sothebys", aid, a.get("title"), parse_date(a.get("dates")),
            cur, "HAMMER")
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_number")),
                "wine": lot.get("title"),
                "estimate_low": to_float(lot.get("estimate_low")),
                "estimate_high": to_float(lot.get("estimate_high")),
                "hammer_price": to_float(lot.get("hammer_price")),
                "realised_price": to_float(lot.get("final_price")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency") or cur,
                "sold": 1 if lot.get("sold") else 0,
            })
        n += 1
        if limit and n >= limit:
            break


def load_christies(b, limit):
    n = 0
    for f in _json_files("Christies"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("event_id") or a.get("sale_id"))
        a_id = b.add_auction(
            "christies", aid, a.get("title"),
            ym_date(a.get("year"), a.get("month")), None, "HAMMER",
            a.get("landing_url"))
        for lot in d.get("lots", []):
            sold = 1 if (lot.get("price_realised") and
                         not lot.get("lot_withdrawn")) else 0
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_number")),
                "wine": lot.get("title"),
                "estimate_low": to_float(lot.get("estimate_low")),
                "estimate_high": to_float(lot.get("estimate_high")),
                "hammer_price": to_float(lot.get("price_realised")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency"),
                "sold": sold,
                "description": lot.get("description"),
                "source_url": lot.get("url"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_zacky(b, limit):
    n = 0
    for f in _json_files("Zacky", "**/auction-*.json"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        cur = a.get("currency")
        a_id = b.add_auction(
            "zacky", aid, a.get("name"), ym_date(a.get("year"), a.get("month")),
            cur, "HAMMER_PLUS_BUYERS_PREMIUM",
            a.get("snapshot_url"), a.get("snapshot_timestamp"))
        for lot in d.get("lots", []):
            fp = lot.get("final_price")
            sold = 1 if (fp and fp > 0) else 0
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": lot.get("name"),
                "quantity": parse_qty_text(lot.get("quantity")),
                "bottle_size_dl": parse_dl(lot.get("quantity")),
                "realised_price": to_float(fp) if fp else None,
                "price_basis": "HAMMER_PLUS_BUYERS_PREMIUM",
                "currency": cur, "sold": sold,
                "description": lot.get("description"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_finarte(b, limit):
    n = 0
    for f in _json_files("Finarte"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("slug"))
        a_id = b.add_auction(
            "finarte", aid, a.get("name"), parse_date(a.get("date")),
            "EUR", "HAMMER")
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_nr")),
                "wine": lot.get("name"), "region": lot.get("region"),
                "vintage": to_int(lot.get("vintage")),
                "estimate_low": to_float(lot.get("estimate_min")),
                "estimate_high": to_float(lot.get("estimate_max")),
                "hammer_price": to_float(lot.get("hammer_price")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency") or "EUR",
                "sold": 1 if lot.get("sold") else 0,
                "condition_text": lot.get("vat_regime"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_dorotheum(b, limit):
    n = 0
    for f in _json_files("Dorotheum"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        a_id = b.add_auction(
            "dorotheum", aid, a.get("name"), parse_date(a.get("date")),
            "EUR", "HAMMER")
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_nr")),
                "wine": lot.get("name"),
                "estimate_low": to_float(lot.get("estimate")),
                "estimate_high": to_float(lot.get("estimate")),
                "hammer_price": to_float(lot.get("realized_price")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency") or "EUR",
                "sold": 1 if lot.get("sold") else 0,
                "description": lot.get("description"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_pandolfini(b, limit):
    n = 0
    for f in _json_files("Pandolfini"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        a_id = b.add_auction(
            "pandolfini", aid, a.get("name"), parse_date(a.get("date")),
            "EUR", "HAMMER_PLUS_BUYERS_PREMIUM")
        for lot in d.get("lots", []):
            sp = lot.get("sold_price") or {}
            sold = 1 if sp.get("class") == "venduto" else 0
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_nr")),
                "wine": lot.get("name"),
                "estimate_low": euro_int(lot.get("estimate_low")),
                "estimate_high": euro_int(lot.get("estimate_high")),
                "realised_price": euro_int(sp.get("value")) if sp else None,
                "price_basis": "HAMMER_PLUS_BUYERS_PREMIUM",
                "currency": "EUR", "sold": sold,
                "description": lot.get("description"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_dobiaschofsky(b, limit):
    n = 0
    for f in _json_files("Dobiaschofsky", "**/auktion-*.json"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        a_id = b.add_auction(
            "dobiaschofsky", aid, a.get("name"), parse_date(a.get("date")),
            "CHF", "HAMMER")
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_nr")),
                "wine": lot.get("name"),
                "vintage": to_int(lot.get("vintage")),
                "estimate_low": to_float(lot.get("estimate_chf")),
                "estimate_high": to_float(lot.get("estimate_chf")),
                "hammer_price": to_float(lot.get("hammer_chf")),
                "price_basis": "HAMMER", "currency": "CHF",
                "description": lot.get("description"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_munich_wine_company(b, limit):
    n = 0
    for f in _json_files("MunichWineCompany"):
        d = load_json(f)
        b.set_source(f)
        a = d.get("auction", {})
        aid = str(a.get("id"))
        cur = a.get("currency") or "EUR"
        a_id = b.add_auction(
            "munich-wine-company", aid, a.get("title"),
            parse_date(a.get("date")), cur, "HAMMER",
            None, a.get("snapshot_timestamp"))
        for lot in d.get("lots", []):
            b.add_lot(a_id, {
                "lot_no": str(lot.get("lot_no")),
                "wine": lot.get("name"), "producer": lot.get("brand") or None,
                "quantity": parse_qty_text(lot.get("count_fill")),
                "hammer_price": to_float(lot.get("hammer_price")),
                "price_basis": "HAMMER",
                "currency": lot.get("currency") or cur,
                "description": lot.get("description"),
                "condition_text": lot.get("count_fill"),
            })
        n += 1
        if limit and n >= limit:
            break


def load_winebarrel(b, limit):
    f = os.path.join(ARCHIVE, "winebarrel", "winebarrel_lots.json")
    d = load_json(f)
    b.set_source(f)
    # one virtual auction holding everything
    a_id = b.add_auction("winebarrel", "all", "winebarrel", None, None, "HAMMER")
    for i, lot in enumerate(d):
        if limit and i >= limit:
            break
        b.add_lot(a_id, {
            "lot_no": str(lot.get("auction_id")),
            "lot_date": parse_date(lot.get("auction_end")),
            "wine": lot.get("name"), "producer": lot.get("winery"),
            "vintage": to_int(lot.get("vintage")),
            "quantity": to_int(lot.get("bottles")),
            "region": lot.get("region"),
            "hammer_price": to_float(lot.get("hammer_price")),
            "price_basis": "HAMMER",
        })


def load_idealwine(b, limit):
    n = 0
    for f in _json_files("IDealwine_normalized"):
        d = load_json(f)
        b.set_source(f)
        aid = "cote"
        a_id = b.add_auction(
            "idealwine", aid, f"iDealwine Cote {d.get('region')}",
            None, "EUR", None)
        for a in d.get("adjudications", []):
            basis = a.get("price_basis")
            b.add_lot(a_id, {
                "lot_no": a.get("code"),
                "lot_date": parse_date(a.get("sold_at")),
                "wine": d.get("wine"), "producer": d.get("estate"),
                "vintage": to_int(d.get("vintage")),
                "region": d.get("region"),
                "classification": d.get("classification") or None,
                "quantity": to_int(a.get("number_of_bottles")),
                "bottle_size_dl": parse_dl(a.get("format")),
                "hammer_price": a.get("hammer_per_bottle_eur")
                or a.get("hammer_per_bottle_eur_derived"),
                "realised_price": a.get("total_per_bottle_eur"),
                "price_basis": basis, "currency": "EUR",
            })
        n += 1
        if limit and n >= limit:
            break


# --- Excel/CSV houses ------------------------------------------------------ #

def load_wermuth(b, limit):
    path = os.path.join(REPO, "VinoImporter", "validatedOutput",
                        "vinoStagingFile2015-2008.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    b.set_source(path)
    n = 0
    for ws in wb.worksheets:
        m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", ws.title)
        date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else None
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        header = [str(c) if c is not None else "" for c in all_rows[0]]
        event = None
        lots = []
        for row in all_rows[1:]:
            rec = dict(zip(header, row))
            if event is None:
                ev = _v(rec, "eventIdentifier")
                if ev and str(ev).strip() not in ("", "null", "None"):
                    event = str(ev)
            lots.append(rec)
        a_id = b.add_auction("wermuth-sa", event or ws.title, ws.title,
                             date, "CHF", "HAMMER")
        for rec in lots:
            rp = _v(rec, "realizedPrice")
            b.add_lot(a_id, {
                "lot_no": str(_v(rec, "providerOfferingId")),
                "wine": _v(rec, "name"),
                "producer": _v(rec, "producer"),
                "vintage": to_int(_v(rec, "vintage")),
                "quantity": to_int(_v(rec, "noOfBottles")),
                "bottle_size_dl": to_float(_v(rec, "deciliters")),
                "estimate_low": to_float(_v(rec, "priceMin")),
                "estimate_high": to_float(_v(rec, "priceMax")),
                "hammer_price": to_float(rp),
                "price_basis": "HAMMER", "currency": "CHF",
                "sold": 1 if rp not in (None, "", "null", "None") else 0,
            })
        n += 1
        if limit and n >= limit:
            break


def _v(rec, key):
    """Header-tolerant field access."""
    if key in rec:
        return rec[key]
    for k in rec:
        if k and k.strip().lower() == key.strip().lower():
            return rec[k]
    return None


def load_steinfels(b, limit):
    base = os.path.join(REPO, "priceData", "import", "steinfels", "prepared")
    files = sorted(glob.glob(os.path.join(base, "**", "results_*.xlsx"),
                             recursive=True))
    n = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        b.set_source(f)
        if "Wines" not in wb.sheetnames:
            continue
        ws = wb["Wines"]
        aid = os.path.basename(f).replace(".xlsx", "").replace("results_", "")
        year = None
        # year from path (e.g. .../2002/results_315.xlsx)
        m = re.search(r"/prepared/(\d{4})/", f)
        if m:
            year = int(m.group(1))
        a_id = b.add_auction("steinfels", aid, aid,
                             f"{year}-01-01" if year else None, "CHF", "HAMMER")
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else "" for c in row]
                continue
            rec = dict(zip(header, row))
            b.add_lot(a_id, {
                "lot_no": str(_v(rec, "Lot")),
                "wine": _v(rec, "Wine"),
                "producer": _v(rec, "Producer") or None,
                "region": _v(rec, "Region") or None,
                "vintage": to_int(_v(rec, "Year")),
                "quantity": to_int(_v(rec, "Quantity")),
                "bottle_size_dl": to_float(_v(rec, "Unit_DL")),
                "hammer_price": to_float(_v(rec, "Price_CHF")),
                "price_basis": "HAMMER", "currency": "CHF",
            })
        n += 1
        if limit and n >= limit:
            break


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

HOUSES = [
    ("baghera", load_baghera),
    ("hdh", load_hdh),
    ("winefields", load_winefields),
    ("besch-cannes", load_beschcannes),
    ("langtons", load_langtons),
    ("sothebys", load_sothebys),
    ("christies", load_christies),
    ("zacky", load_zacky),
    ("finarte", load_finarte),
    ("dorotheum", load_dorotheum),
    ("pandolfini", load_pandolfini),
    ("dobiaschofsky", load_dobiaschofsky),
    ("munich-wine-company", load_munich_wine_company),
    ("winebarrel", load_winebarrel),
    ("idealwine", load_idealwine),
    ("wermuth-sa", load_wermuth),
    ("steinfels", load_steinfels),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--only", type=str, default=None,
                    help="comma-separated house names to include")
    ap.add_argument("--skip-resolution", action="store_true",
                    help="load raw lots only, skip Stufe 2 resolution")
    args = ap.parse_args()

    if args.dry_run:
        conn = sqlite3.connect(":memory:")
        conn.execute("PRAGMA foreign_keys = ON")
        _build(conn, args)
        return

    lock_path = os.path.join(SCRIPT_DIR, ".wine-db-build.lock")
    lock_fd = _acquire_lock(lock_path)
    if lock_fd is None:
        print("Build already running (lock held).", file=sys.stderr)
        sys.exit(1)
    tmp_path = f"{DB_PATH}.tmp.{os.getpid()}"
    try:
        conn = sqlite3.connect(tmp_path)
        conn.execute("PRAGMA foreign_keys = ON")
        _build(conn, args)
        _validate_gates(conn)
        conn.commit()
        conn.close()
        os.replace(tmp_path, DB_PATH)
        print(f"published: {DB_PATH}")
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    finally:
        os.close(lock_fd)
        if os.path.exists(lock_path):
            os.remove(lock_path)


def _build(conn, args):
    conn.executescript(SCHEMA)
    b = Builder(conn)

    totals = {}
    houses = HOUSES
    if args.only:
        wanted = {h.strip() for h in args.only.split(",") if h.strip()}
        houses = [(n, f) for n, f in HOUSES if n in wanted]
    for name, fn in houses:
        before = conn.execute("SELECT count(*) FROM lots").fetchone()[0]
        try:
            fn(b, args.limit)
        except Exception as e:
            print(f"ERROR {name}: {type(e).__name__}: {e}", file=sys.stderr)
            raise
        after = conn.execute("SELECT count(*) FROM lots").fetchone()[0]
        totals[name] = after - before
        print(f"  {name:22} +{after - before} lots")

    na = conn.execute("SELECT count(*) FROM auctions").fetchone()[0]
    nl = conn.execute("SELECT count(*) FROM lots").fetchone()[0]
    np_ = conn.execute("SELECT count(*) FROM providers").fetchone()[0]
    print(f"\nproviders={np_} auctions={na} lots={nl}")

    if not args.skip_resolution:
        from build_wines import run as resolve_wines
        resolve_wines(conn)

    _record_build_metadata(conn)


def _acquire_lock(lock_path):
    try:
        fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(os.getpid()).encode())
        return fd
    except FileExistsError:
        return None


def _validate_gates(conn):
    problems = []
    ic = conn.execute("PRAGMA integrity_check").fetchone()
    if ic and ic[0] != "ok":
        problems.append(f"integrity_check: {ic[0]}")
    fk = conn.execute("PRAGMA foreign_key_check").fetchall()
    if fk:
        problems.append(f"foreign_key_check: {len(fk)} violations")
    mixed_ref = conn.execute(
        "SELECT count(*) FROM lots WHERE lot_kind='MIXED'"
        " AND wine_ref_id IS NOT NULL").fetchone()[0]
    if mixed_ref:
        problems.append(f"mixed lots with wine_ref_id: {mixed_ref}")
    conflict_final = conn.execute(
        "SELECT count(*) FROM lots WHERE vintage_status='CONFLICT'"
        " AND vintage_final IS NOT NULL").fetchone()[0]
    if conflict_final:
        problems.append(f"conflict lots with vintage_final: {conflict_final}")
    if problems:
        raise RuntimeError("validation gates failed: " + "; ".join(problems))


def _record_build_metadata(conn):
    import hashlib
    import platform

    from wine_resolution import (LOADER_VERSION, NORMALIZER_VERSION,
                                 RESOLVER_VERSION)

    files = sorted(glob.glob(os.path.join(ARCHIVE, "**", "*.json"),
                             recursive=True))
    files += sorted(glob.glob(os.path.join(REPO, "VinoImporter",
                                           "validatedOutput", "*.xlsx")))
    files += sorted(glob.glob(os.path.join(REPO, "priceData", "import",
                                           "steinfels", "prepared", "**",
                                           "*.xlsx"), recursive=True))
    files = [f for f in files if "IDealwine_normalized" not in f
             and "wines_directory" not in f and "auctions_directory" not in f]

    build_id = hashlib.sha256(
        repr(sorted(files)).encode("utf-8")).hexdigest()[:32]
    entries = []
    for f in files:
        with open(f, "rb") as fh:
            sha = hashlib.sha256(fh.read()).hexdigest()
        size = os.path.getsize(f)
        entries.append((os.path.relpath(f, ARCHIVE), sha, size))
    entries.sort()
    manifest_sha = hashlib.sha256(repr(entries).encode("utf-8")).hexdigest()

    now = __import__("datetime").datetime.now().isoformat()
    conn.execute(
        "INSERT INTO builds(id, started_at, completed_at,"
        " input_manifest_sha256, loader_version, normalizer_version,"
        " resolver_version, python_version, sqlite_version, status)"
        " VALUES (?,?,?,?,?,?,?,?,?,?)",
        (build_id, now, now, manifest_sha, LOADER_VERSION, NORMALIZER_VERSION,
         RESOLVER_VERSION, platform.python_version(),
         sqlite3.sqlite_version, "OK"))
    for rel, sha, size in entries:
        conn.execute(
            "INSERT INTO build_inputs(build_id, source_file,"
            " source_file_sha256, file_size) VALUES (?,?,?,?)",
            (build_id, rel, sha, size))


if __name__ == "__main__":
    main()
